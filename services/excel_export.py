"""
Exportação para Excel (.xlsx) formatada: cabeçalho, filtros, colunas ajustadas,
datas formatadas, cabeçalho congelado.
"""
import os
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from database.db import get_connection

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _formatar_aba(ws, n_colunas):
    for col in range(1, n_colunas + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_colunas)}1"
    for col in range(1, n_colunas + 1):
        letra = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letra].width = min(max(max_len + 2, 10), 50)


def _query_requisicoes(data_inicio: str, data_fim: str, filtros: dict = None):
    filtros = filtros or {}
    conn = get_connection()
    sql = """
        SELECT r.numero AS requisicao, r.data, r.hora,
               p.codigo, p.nome AS produto, ri.quantidade, p.unidade,
               f.nome AS solicitante, f.funcao,
               v.placa AS veiculo,
               u.nome AS almoxarife
        FROM requisicoes r
        JOIN requisicao_itens ri ON ri.requisicao_id = r.id
        JOIN produtos p ON p.id = ri.produto_id
        JOIN funcionarios f ON f.id = r.funcionario_id
        JOIN veiculos v ON v.id = r.veiculo_id
        JOIN usuarios u ON u.id = r.almoxarife_id
        WHERE r.data BETWEEN ? AND ? AND r.status = 'Confirmada'
    """
    params = [data_inicio, data_fim]
    if filtros.get("solicitante"):
        sql += " AND f.nome = ?"
        params.append(filtros["solicitante"])
    if filtros.get("veiculo"):
        sql += " AND v.placa = ?"
        params.append(filtros["veiculo"])
    if filtros.get("produto"):
        sql += " AND p.nome LIKE ?"
        params.append(f"%{filtros['produto']}%")
    if filtros.get("almoxarife"):
        sql += " AND u.nome = ?"
        params.append(filtros["almoxarife"])
    sql += " ORDER BY r.data, r.hora, r.numero"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


def exportar_requisicoes(data_inicio: str, data_fim: str, nome_arquivo: str, filtros: dict = None):
    rows = _query_requisicoes(data_inicio, data_fim, filtros)
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requisições"
    colunas = ["Requisição", "Data", "Hora", "Código", "Produto", "Quantidade",
               "Unidade", "Solicitante", "Função", "Veículo", "Almoxarife"]
    ws.append(colunas)
    for r in rows:
        ws.append([r["requisicao"], r["data"], r["hora"], r["codigo"], r["produto"],
                   r["quantidade"], r["unidade"], r["solicitante"], r["funcao"],
                   r["veiculo"], r["almoxarife"]])
    _formatar_aba(ws, len(colunas))

    # aba resumo
    ws2 = wb.create_sheet("Resumo por Solicitante")
    ws2.append(["Solicitante", "Requisições", "Itens Retirados"])
    resumo = {}
    for r in rows:
        chave = r["solicitante"]
        resumo.setdefault(chave, {"reqs": set(), "itens": 0})
        resumo[chave]["reqs"].add(r["requisicao"])
        resumo[chave]["itens"] += r["quantidade"]
    for solicitante, d in sorted(resumo.items()):
        ws2.append([solicitante, len(d["reqs"]), d["itens"]])
    _formatar_aba(ws2, 3)

    ws3 = wb.create_sheet("Resumo por Veículo")
    ws3.append(["Veículo", "Requisições", "Itens Retirados"])
    resumo_v = {}
    for r in rows:
        chave = r["veiculo"]
        resumo_v.setdefault(chave, {"reqs": set(), "itens": 0})
        resumo_v[chave]["reqs"].add(r["requisicao"])
        resumo_v[chave]["itens"] += r["quantidade"]
    for veiculo, d in sorted(resumo_v.items()):
        ws3.append([veiculo, len(d["reqs"]), d["itens"]])
    _formatar_aba(ws3, 3)

    path = os.path.join(EXPORTS_DIR, nome_arquivo)
    wb.save(path)
    return path, len(rows)


def exportar_diario(data: str):
    nome = f"Requisicoes_Diarias_{data.replace('-', '_')}.xlsx"
    return exportar_requisicoes(data, data, nome)


def exportar_semanal(data_inicio: str, data_fim: str):
    nome = f"Requisicoes_Semanais_{data_inicio.replace('-', '_')}_a_{data_fim.replace('-', '_')}.xlsx"
    return exportar_requisicoes(data_inicio, data_fim, nome)


def exportar_mensal(ano: int, mes: int):
    data_inicio = f"{ano}-{mes:02d}-01"
    if mes == 12:
        prox = dt.date(ano + 1, 1, 1)
    else:
        prox = dt.date(ano, mes + 1, 1)
    data_fim = (prox - dt.timedelta(days=1)).strftime("%Y-%m-%d")
    meses_pt = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho",
                "Agosto","Setembro","Outubro","Novembro","Dezembro"]
    nome = f"Requisicoes_Mensais_{meses_pt[mes-1]}-{ano}.xlsx"
    return exportar_requisicoes(data_inicio, data_fim, nome)


def exportar_estoque_atual():
    conn = get_connection()
    rows = conn.execute(
        "SELECT codigo, nome, secao, unidade, estoque_atual, estoque_minimo, estoque_maximo, status "
        "FROM produtos ORDER BY secao, nome"
    ).fetchall()
    conn.close()

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"
    colunas = ["Código", "Nome", "Seção", "Unidade", "Estoque Atual", "Estoque Mínimo", "Estoque Máximo", "Status"]
    ws.append(colunas)
    for r in rows:
        ws.append([r["codigo"], r["nome"], r["secao"], r["unidade"], r["estoque_atual"],
                   r["estoque_minimo"], r["estoque_maximo"], r["status"]])
    _formatar_aba(ws, len(colunas))

    hoje = dt.date.today().strftime("%d-%m-%Y")
    nome_arquivo = f"Estoque_Atual_{hoje}.xlsx"
    path = os.path.join(EXPORTS_DIR, nome_arquivo)
    wb.save(path)
    return path, len(rows)


def exportar_faturamento_diario(data: str):
    """Export simplificado para o setor de faturamento: uma linha por item retirado no dia."""
    rows = _query_requisicoes(data, data)
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Retiradas do Dia"
    colunas = ["Data", "Requisição", "Código", "Produto", "Quantidade", "Veículo", "Solicitante"]
    ws.append(colunas)
    for r in rows:
        data_fmt = dt.datetime.strptime(r["data"], "%Y-%m-%d").strftime("%d/%m/%Y")
        ws.append([data_fmt, r["requisicao"], r["codigo"], r["produto"], r["quantidade"],
                   r["veiculo"], r["solicitante"]])
    _formatar_aba(ws, len(colunas))
    nome_arquivo = f"Faturamento_{data.replace('-', '_')}.xlsx"
    path = os.path.join(EXPORTS_DIR, nome_arquivo)
    wb.save(path)
    return path, len(rows)
